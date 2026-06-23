"""Generate per-action confusion matrix + frame axiom violation spreadsheets.

Compares gtrate0 (noisy) trajectories to gtrate100 (GT) trajectories for a given
planning domain, producing an Excel workbook with:
  - One sheet per action: TP/FP/FN/TN per lifted fluent per problem
  - One sheet for frame axiom violations: per-transition FA detail

Usage:
    python generate_trajectory_confusion_spreadsheet.py \\
        --base <trajectory_dir> \\
        --domain-pddl <path_to_domain.pddl> \\
        --output <output.xlsx>

Or import and call `process_domain(...)` directly.
"""

import re
import argparse
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# Trajectory parsing
# =============================================================================

def parse_trajectory(path: Path):
    """Parse a .trajectory file into (list[set[str]], list[str]) = (states, actions)."""
    text = path.read_text()
    states, actions = [], []
    for line in text.strip().split("\n"):
        line = line.strip()
        if line.startswith("(:init") or line.startswith("(:state"):
            tag = "(:init" if line.startswith("(:init") else "(:state"
            content = line[len(tag):]
            if content.endswith(")"):
                content = content[:-1]  # strip only the closing paren of (:state/(:init
            fluents = set()
            for m in re.finditer(r"\(([^)]+)\)", content):
                fluents.add(m.group(1).strip())
            states.append(fluents)
        elif line.startswith("(operator:"):
            actions.append(re.search(r"\(operator:\s*\(([^)]+)\)\)", line).group(1))
    return states, actions


# =============================================================================
# Domain model parsing
# =============================================================================

def _find_balanced(text: str, start: int) -> int:
    """Return the index of the closing paren matching the '(' at *start*."""
    depth = 0
    for i in range(start, len(text)):
        if text[i] == "(":
            depth += 1
        elif text[i] == ")":
            depth -= 1
            if depth == 0:
                return i
    return -1


def _extract_top_level_sexps(text: str) -> list:
    """Return a list of top-level S-expressions (as strings) from *text*."""
    result = []
    i = 0
    while i < len(text):
        if text[i] == "(":
            j = _find_balanced(text, i)
            if j == -1:
                break
            result.append(text[i : j + 1])
            i = j + 1
        else:
            i += 1
    return result


def parse_domain_pddl(pddl_path: Path) -> dict:
    """Parse a PDDL domain file into an action model dict.

    Returns:
        {action_name: {params: [role_names], preconditions: [(pred, [roles])],
                       add_effects: [...], del_effects: [...]}}
    """
    text = pddl_path.read_text()

    action_model = {}

    # Find each (:action ...) block using balanced parens
    idx = 0
    while True:
        pos = text.find("(:action", idx)
        if pos == -1:
            break
        end = _find_balanced(text, pos)
        if end == -1:
            break
        block = text[pos : end + 1]
        idx = end + 1

        # Extract action name
        name_match = re.match(r"\(:action\s+(\S+)", block)
        if not name_match:
            continue
        action_name = name_match.group(1)

        # Extract :parameters section
        params_pos = block.find(":parameters")
        if params_pos == -1:
            continue
        paren_start = block.index("(", params_pos + len(":parameters"))
        paren_end = _find_balanced(block, paren_start)
        params_str = block[paren_start + 1 : paren_end]
        params = [m.group(1) for m in re.finditer(r"\?(\w+)", params_str)]

        # Extract :precondition section
        def extract_section(keyword):
            kpos = block.find(keyword)
            if kpos == -1:
                return ""
            pstart = block.index("(", kpos + len(keyword))
            pend = _find_balanced(block, pstart)
            return block[pstart : pend + 1]

        pre_block = extract_section(":precondition")
        eff_block = extract_section(":effect")

        # Parse literals from a block like (and (lit1) (not (lit2)) ...)
        def parse_literals(s):
            pos_lits, neg_lits = [], []
            # Strip outer (and ...) if present
            inner = s.strip()
            if inner.startswith("(and"):
                inner = inner[4:-1].strip()  # remove (and ... )
            elif inner.startswith("("):
                inner = inner[1:-1].strip()  # remove outer parens of single literal

            for sexp in _extract_top_level_sexps(inner):
                content = sexp[1:-1].strip()  # strip outer parens
                if content.startswith("not "):
                    # Negative literal: (not (pred ?args...))
                    inner_sexp = content[4:].strip()
                    inner_content = inner_sexp[1:-1].strip()  # strip inner parens
                    parts = inner_content.split()
                    pred = parts[0]
                    roles = [p.lstrip("?") for p in parts[1:]]
                    neg_lits.append((pred, roles))
                else:
                    parts = content.split()
                    pred = parts[0]
                    if pred == "and":
                        continue
                    roles = [p.lstrip("?") for p in parts[1:]]
                    pos_lits.append((pred, roles))
            return pos_lits, neg_lits

        pre_pos, _ = parse_literals(pre_block)
        eff_pos, eff_neg = parse_literals(eff_block)

        # Normalize hyphens to underscores (trajectory files use underscores)
        action_name_norm = action_name.replace("-", "_")

        action_model[action_name_norm] = {
            "params": params,
            "preconditions": pre_pos,
            "add_effects": eff_pos,
            "del_effects": eff_neg,
        }

    return action_model


# =============================================================================
# Confusion matrix computation
# =============================================================================

def ground_fluent(pred_name: str, roles: list, bindings: dict) -> str:
    args = [bindings[r] for r in roles]
    return f"{pred_name} {' '.join(args)}" if args else pred_name


def _has_complementary_swap(pred_name: str, ground_args: list, error_type: str,
                            noisy_state: set, gt_state: set) -> bool:
    """Check if a complementary swap exists for this binary predicate grounding.

    A swap occurs when this grounding has an error (FP or FN) and another
    grounding of the same predicate — sharing at least one argument — has the
    complementary error type.
    """
    fp_fluents = noisy_state - gt_state   # present in noisy, absent in GT
    fn_fluents = gt_state - noisy_state   # present in GT, absent in noisy
    target_set = fn_fluents if error_type == "FP" else fp_fluents

    ground_args_set = set(ground_args)
    for f in target_set:
        f_parts = f.split()
        if f_parts[0] != pred_name:
            continue
        f_args = f_parts[1:]
        if set(f_args) & ground_args_set:  # shares at least one argument
            return True
    return False


def compute_confusion(noisy_states, gt_states, actions, action_model):
    result = defaultdict(
        lambda: defaultdict(lambda: defaultdict(
            lambda: {"TP": 0, "FP": 0, "FN": 0, "TN": 0, "SW": 0}))
    )
    action_counts = defaultdict(int)

    for i, action_str in enumerate(actions):
        parts = action_str.split()
        action_name = parts[0]
        action_args = parts[1:]
        action_counts[action_name] += 1
        model = action_model[action_name]
        bindings = dict(zip(model["params"], action_args))

        prev_n, prev_g = noisy_states[i], gt_states[i]
        next_n, next_g = noisy_states[i + 1], gt_states[i + 1]

        def _classify_and_record(cat, pred_name, roles, noisy_state, gt_state):
            g = ground_fluent(pred_name, roles, bindings)
            lifted = f"{pred_name}({','.join(roles)})"
            in_n, in_g = g in noisy_state, g in gt_state
            is_binary = len(roles) >= 2

            if cat == "DEL":
                if not in_g and not in_n:
                    result[action_name][cat][lifted]["TN"] += 1
                elif in_n and not in_g:
                    result[action_name][cat][lifted]["FP"] += 1
                    if is_binary:
                        g_args = g.split()[1:]
                        if _has_complementary_swap(pred_name, g_args, "FP", noisy_state, gt_state):
                            result[action_name][cat][lifted]["SW"] += 1
                elif in_g and not in_n:
                    result[action_name][cat][lifted]["FN"] += 1
                    if is_binary:
                        g_args = g.split()[1:]
                        if _has_complementary_swap(pred_name, g_args, "FN", noisy_state, gt_state):
                            result[action_name][cat][lifted]["SW"] += 1
                else:
                    result[action_name][cat][lifted]["TP"] += 1
            else:
                if in_g and in_n:
                    result[action_name][cat][lifted]["TP"] += 1
                elif in_n:
                    result[action_name][cat][lifted]["FP"] += 1
                    if is_binary:
                        g_args = g.split()[1:]
                        if _has_complementary_swap(pred_name, g_args, "FP", noisy_state, gt_state):
                            result[action_name][cat][lifted]["SW"] += 1
                elif in_g:
                    result[action_name][cat][lifted]["FN"] += 1
                    if is_binary:
                        g_args = g.split()[1:]
                        if _has_complementary_swap(pred_name, g_args, "FN", noisy_state, gt_state):
                            result[action_name][cat][lifted]["SW"] += 1
                else:
                    result[action_name][cat][lifted]["TN"] += 1

        # Preconditions: check prev_state
        for pred_name, roles in model["preconditions"]:
            _classify_and_record("PRE", pred_name, roles, prev_n, prev_g)

        # Add effects: check next_state
        for pred_name, roles in model["add_effects"]:
            _classify_and_record("ADD", pred_name, roles, next_n, next_g)

        # Del effects: check next_state
        for pred_name, roles in model["del_effects"]:
            _classify_and_record("DEL", pred_name, roles, next_n, next_g)

    return result, action_counts


# =============================================================================
# Frame axiom violation computation
# =============================================================================

def get_fluent_objects(fluent_str: str) -> set:
    parts = fluent_str.split()
    return set(parts[1:])


def compute_fa_violations(noisy_states, gt_states, actions):
    results = []
    for i, action_str in enumerate(actions):
        parts = action_str.split()
        action_objs = set(parts[1:])
        next_n, next_g = noisy_states[i + 1], gt_states[i + 1]
        num_preds_gt = len(next_g)
        fp_f = next_n - next_g
        fn_f = next_g - next_n
        wrong = fp_f | fn_f
        fa = []
        for f in sorted(wrong):
            f_objs = get_fluent_objects(f)
            if len(f_objs) == 0:
                continue
            if not (f_objs <= action_objs):
                fa.append({"fluent": f, "type": "FP" if f in fp_f else "FN"})
        results.append(
            {
                "transition": i,
                "action": action_str,
                "total_errors_in_next": len(wrong),
                "fa_violations": fa,
                "fa_count": len(fa),
                "num_preds": num_preds_gt,
            }
        )
    return results


# =============================================================================
# Excel generation
# =============================================================================

# Shared styles
_HEADER_FONT_WHITE = Font(name="Arial", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
_DATA_FONT = Font(name="Arial", size=10)
_FA_FILL = PatternFill("solid", fgColor="FCE4EC")
_CLEAN_FILL = PatternFill("solid", fgColor="E8F5E9")
_SUMMARY_FILL = PatternFill("solid", fgColor="FFF3E0")
_WARN_FILL = PatternFill("solid", fgColor="FFF9C4")
_NA_FILL = PatternFill("solid", fgColor="D9D9D9")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_cell(ws, ref, font=None, fill=None, alignment=None, border=None):
    cell = ws[ref]
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def build_action_sheet(ws, action_name, model, problems, per_problem_confusion, per_problem_action_counts):
    """Build a confusion-matrix sheet for one action."""
    ws.merge_cells("A1:L1")
    ws["A1"] = f"ACTION: {action_name}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)

    row = 4
    ws[f"A{row}"] = "Cat"
    ws[f"B{row}"] = "GT Fluent"
    for c in ["A", "B"]:
        _style_cell(ws, f"{c}{row}", _HEADER_FONT_WHITE, _HEADER_FILL, _CENTER, _THIN_BORDER)

    for pi, prob in enumerate(problems):
        if prob not in per_problem_confusion:
            continue
        cl = chr(ord("C") + pi)
        n = per_problem_action_counts.get(prob, {}).get(action_name, 0)
        ws[f"{cl}{row}"] = f"{prob}\n(n={n})"
        _style_cell(ws, f"{cl}{row}", _HEADER_FONT_WHITE, _HEADER_FILL, _CENTER, _THIN_BORDER)

    row = 5
    sections = [
        ("PRECONDITIONS (in prev state)", "PRE", model["preconditions"]),
        ("ADD EFFECTS (in next state)", "ADD", model["add_effects"]),
        ("DELETE EFFECTS (absent from next state)", "DEL", model["del_effects"]),
    ]

    for section_label, cat, fluent_list in sections:
        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = section_label
        ws[f"A{row}"].font = Font(name="Arial", bold=True, size=10)
        ws[f"A{row}"].fill = _SECTION_FILL
        row += 1

        for pred_name, roles in fluent_list:
            lifted = f"{pred_name}({','.join(roles)})"
            ws[f"A{row}"] = cat
            ws[f"B{row}"] = lifted
            for c in ["A", "B"]:
                _style_cell(ws, f"{c}{row}", _DATA_FONT, None, _CENTER, _THIN_BORDER)

            is_binary = len(roles) >= 2
            for pi, prob in enumerate(problems):
                if prob not in per_problem_confusion:
                    continue
                cl = chr(ord("C") + pi)
                n = per_problem_action_counts.get(prob, {}).get(action_name, 0)
                if n == 0:
                    ws[f"{cl}{row}"] = "N/A"
                    _style_cell(ws, f"{cl}{row}", _DATA_FONT, _NA_FILL, _CENTER, _THIN_BORDER)
                else:
                    cm = per_problem_confusion[prob][action_name][cat][lifted]
                    cell_text = f"TP={cm['TP']} FP={cm['FP']}\nFN={cm['FN']} TN={cm['TN']}"
                    if is_binary:
                        cell_text += f"\nSW={cm['SW']}"
                    ws[f"{cl}{row}"] = cell_text
                    ws[f"{cl}{row}"].fill = _WARN_FILL if (cm["FP"] > 0 or cm["FN"] > 0) else _CLEAN_FILL
                    _style_cell(ws, f"{cl}{row}", _DATA_FONT, None, _CENTER, _THIN_BORDER)
            row += 1
        row += 1  # blank row between sections

    # N transitions row
    ws[f"B{row}"] = "N transitions"
    ws[f"B{row}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"B{row}"].border = _THIN_BORDER
    for pi, prob in enumerate(problems):
        if prob not in per_problem_confusion:
            continue
        cl = chr(ord("C") + pi)
        n = per_problem_action_counts.get(prob, {}).get(action_name, 0)
        ws[f"{cl}{row}"] = n
        ws[f"{cl}{row}"].font = Font(name="Arial", bold=True, size=10)
        ws[f"{cl}{row}"].alignment = _CENTER
        ws[f"{cl}{row}"].border = _THIN_BORDER
        if n == 0:
            ws[f"{cl}{row}"].fill = _NA_FILL

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    for pi in range(len(problems)):
        ws.column_dimensions[chr(ord("C") + pi)].width = 16


def build_fa_sheet(ws, domain_title, problems, all_fa_results):
    """Build the frame-axiom-violations sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{domain_title} — FRAME AXIOM VIOLATIONS: gtrate0 vs GT"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "FA violation = fluent differing from GT in next_state, whose objects "
        "are NOT a subset of action params. Predicate counts from GT next_state."
    )
    ws["A2"].font = Font(name="Arial", size=9, italic=True)

    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "SUMMARY PER PROBLEM"
    ws[f"A{row}"].font = Font(name="Arial", bold=True, size=12)
    row += 1

    for cl, h in [
        ("A", "Problem"),
        ("B", "N transitions"),
        ("C", "P predicates\n(total across\nnext_states)"),
        ("D", "Total errors\n(vs GT)"),
        ("E", "FA violations"),
        ("F", "Non-FA errors"),
        ("G", "FA %"),
        ("H", "Non-FA %"),
    ]:
        ws[f"{cl}{row}"] = h
        _style_cell(ws, f"{cl}{row}", _HEADER_FONT_WHITE, _HEADER_FILL, _CENTER, _THIN_BORDER)
    row += 1

    summary_start = row
    for prob in problems:
        if prob not in all_fa_results:
            continue
        results = all_fa_results[prob]
        nt = len(results)
        tp = sum(r["num_preds"] for r in results)
        te = sum(r["total_errors_in_next"] for r in results)
        tf = sum(r["fa_count"] for r in results)
        nf = te - tf
        fa_pct = tf / te * 100 if te > 0 else 0
        nfa_pct = nf / te * 100 if te > 0 else 0

        ws[f"A{row}"] = prob
        ws[f"A{row}"].font = Font(name="Arial", bold=True, size=10)
        ws[f"B{row}"] = nt
        ws[f"C{row}"] = tp
        ws[f"D{row}"] = te
        ws[f"E{row}"] = tf
        ws[f"F{row}"] = nf
        ws[f"G{row}"] = f"{fa_pct:.1f}%"
        ws[f"H{row}"] = f"{nfa_pct:.1f}%"
        for c in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            _style_cell(ws, f"{c}{row}", _DATA_FONT, None, _CENTER, _THIN_BORDER)
            if tf > 10:
                ws[f"{c}{row}"].fill = _FA_FILL
            elif tf == 0:
                ws[f"{c}{row}"].fill = _CLEAN_FILL
        row += 1

    # Totals row
    ws[f"A{row}"] = "TOTAL"
    ws[f"A{row}"].font = Font(name="Arial", bold=True, size=10)
    for c in ["B", "C", "D", "E", "F"]:
        ws[f"{c}{row}"] = f"=SUM({c}{summary_start}:{c}{row - 1})"
        ws[f"{c}{row}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"G{row}"] = f'=IF(D{row}>0,TEXT(E{row}/D{row}*100,"0.0")&"%","0%")'
    ws[f"G{row}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"H{row}"] = f'=IF(D{row}>0,TEXT(F{row}/D{row}*100,"0.0")&"%","0%")'
    ws[f"H{row}"].font = Font(name="Arial", bold=True, size=10)
    for c in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws[f"{c}{row}"].fill = _SUMMARY_FILL
        ws[f"{c}{row}"].alignment = _CENTER
        ws[f"{c}{row}"].border = _THIN_BORDER

    row += 3

    # Detailed per-problem tables
    for prob in problems:
        if prob not in all_fa_results:
            continue
        results = all_fa_results[prob]
        tf = sum(r["fa_count"] for r in results)
        tp = sum(r["num_preds"] for r in results)

        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = (
            f"{prob.upper()} — {len(results)} transitions, "
            f"{tp} predicates, {tf} FA violations"
        )
        ws[f"A{row}"].font = Font(name="Arial", bold=True, size=11)
        row += 1

        for cl, h in [
            ("A", "Transition"),
            ("B", "Action"),
            ("C", "Predicates\n(next_state)"),
            ("D", "Errors\nvs GT"),
            ("E", "FA\nviolations"),
            ("F", "Violating fluents"),
            ("G", "Type"),
        ]:
            ws[f"{cl}{row}"] = h
            _style_cell(ws, f"{cl}{row}", _HEADER_FONT_WHITE, _HEADER_FILL, _CENTER, _THIN_BORDER)
        row += 1

        for r in results:
            if r["fa_count"] == 0:
                ws[f"A{row}"] = f"T{r['transition']}"
                ws[f"B{row}"] = r["action"]
                ws[f"C{row}"] = r["num_preds"]
                ws[f"D{row}"] = r["total_errors_in_next"]
                ws[f"E{row}"] = 0
                ws[f"F{row}"] = "—"
                ws[f"G{row}"] = "—"
                for c in ["A", "B", "C", "D", "E", "F", "G"]:
                    _style_cell(ws, f"{c}{row}", _DATA_FONT, _CLEAN_FILL, _CENTER, _THIN_BORDER)
                row += 1
            else:
                first_row = row
                for v in r["fa_violations"]:
                    ws[f"F{row}"] = v["fluent"]
                    ws[f"G{row}"] = v["type"]
                    for c in ["F", "G"]:
                        _style_cell(ws, f"{c}{row}", _DATA_FONT, _FA_FILL, _CENTER, _THIN_BORDER)
                    row += 1
                last_row = row - 1
                if first_row < last_row:
                    for c in ["A", "B", "C", "D", "E"]:
                        ws.merge_cells(f"{c}{first_row}:{c}{last_row}")
                ws[f"A{first_row}"] = f"T{r['transition']}"
                ws[f"B{first_row}"] = r["action"]
                ws[f"C{first_row}"] = r["num_preds"]
                ws[f"D{first_row}"] = r["total_errors_in_next"]
                ws[f"E{first_row}"] = r["fa_count"]
                for c in ["A", "B", "C", "D", "E"]:
                    _style_cell(ws, f"{c}{first_row}", _DATA_FONT, _FA_FILL, _CENTER, _THIN_BORDER)
        row += 2

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 10


# =============================================================================
# GT state simulation from action model
# =============================================================================

def simulate_gt_states(init_state: set, actions: list, action_model: dict) -> list:
    """Simulate ground-truth states by applying action effects from init_state.

    Args:
        init_state: Set of fluent strings for the initial state.
        actions: List of grounded action strings (e.g. "move_up player-1 loc-5-4 loc-5-5").
        action_model: Dict from parse_domain_pddl().

    Returns:
        List of N+1 state sets: [init, after_action_0, after_action_1, ...].
    """
    states = [init_state]
    current = set(init_state)

    for action_str in actions:
        parts = action_str.split()
        action_name = parts[0]
        action_args = parts[1:]
        model = action_model[action_name]
        bindings = dict(zip(model["params"], action_args))

        # Apply delete effects first, then add effects
        for pred_name, roles in model["del_effects"]:
            g = ground_fluent(pred_name, roles, bindings)
            current.discard(g)

        for pred_name, roles in model["add_effects"]:
            g = ground_fluent(pred_name, roles, bindings)
            current.add(g)

        states.append(set(current))

    return states


# =============================================================================
# Main pipeline
# =============================================================================

def process_domain(base_path: Path, output_path: Path, action_model: dict,
                   domain_title: str, single_trajectory: bool = False):
    """End-to-end: parse trajectories, compute metrics, write Excel.

    Args:
        base_path: Directory containing problem subdirs.
        output_path: Where to write the .xlsx file.
        action_model: Dict from parse_domain_pddl() or hand-coded.
        domain_title: Human-readable domain name for sheet headers.
        single_trajectory: If True, expect a single ``problemN.trajectory``
            per problem dir (noisy only). GT states are simulated from the
            init state using the action model.
    """
    problems = sorted(
        [d.name for d in base_path.iterdir() if d.is_dir()],
        key=lambda x: int(x.replace("problem", "")),
    )

    per_problem_confusion = {}
    per_problem_action_counts = {}
    all_fa_results = {}

    for prob in problems:
        if single_trajectory:
            noisy_path = base_path / prob / f"{prob}.trajectory"
            if not noisy_path.exists():
                continue
            noisy_states, noisy_actions = parse_trajectory(noisy_path)
            gt_states = simulate_gt_states(noisy_states[0], noisy_actions, action_model)
        else:
            noisy_path = base_path / prob / f"{prob}_gtrate0_frame_axioms.trajectory"
            gt_path = base_path / prob / f"{prob}_gtrate100_frame_axioms.trajectory"
            if not noisy_path.exists() or not gt_path.exists():
                continue
            noisy_states, noisy_actions = parse_trajectory(noisy_path)
            gt_states, gt_actions = parse_trajectory(gt_path)
            if noisy_actions != gt_actions:
                print(f"WARNING: action mismatch in {prob}, skipping")
                continue

        confusion, counts = compute_confusion(noisy_states, gt_states, noisy_actions, action_model)
        per_problem_confusion[prob] = confusion
        per_problem_action_counts[prob] = counts
        all_fa_results[prob] = compute_fa_violations(noisy_states, gt_states, noisy_actions)

    # Build workbook
    wb = Workbook()
    del wb[wb.sheetnames[0]]  # remove default sheet

    action_order = list(action_model.keys())
    for action_name in action_order:
        ws = wb.create_sheet(action_name)
        build_action_sheet(
            ws, action_name, action_model[action_name],
            problems, per_problem_confusion, per_problem_action_counts,
        )

    ws = wb.create_sheet("frame_axiom_violations")
    build_fa_sheet(ws, domain_title, problems, all_fa_results)

    wb.save(output_path)
    print(f"Saved {output_path} with sheets: {wb.sheetnames}")

    # Print summary
    for prob in problems:
        if prob not in all_fa_results:
            continue
        r = all_fa_results[prob]
        tf = sum(x["fa_count"] for x in r)
        te = sum(x["total_errors_in_next"] for x in r)
        print(f"  {prob}: {len(r)} transitions, {te} errors, {tf} FA violations")


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate trajectory confusion matrix spreadsheet"
    )
    parser.add_argument(
        "--base", type=Path, required=True,
        help="Path to trajectories dir (contains problem0/, problem1/, ...)",
    )
    parser.add_argument(
        "--domain-pddl", type=Path, required=True,
        help="Path to the GT domain PDDL file",
    )
    parser.add_argument(
        "--output", type=Path, required=True,
        help="Output .xlsx file path",
    )
    parser.add_argument(
        "--title", type=str, default=None,
        help="Domain title for sheet headers (default: inferred from PDDL filename)",
    )
    parser.add_argument(
        "--single-trajectory", action="store_true",
        help="Each problem dir has a single problemN.trajectory (noisy only). "
             "GT states are simulated from the init state using the domain model.",
    )
    args = parser.parse_args()

    action_model = parse_domain_pddl(args.domain_pddl)
    if not action_model:
        print(f"ERROR: No actions found in {args.domain_pddl}")
        return

    title = args.title or args.domain_pddl.stem.upper().replace("_", " ")
    print(f"Domain: {title}")
    print(f"Actions: {list(action_model.keys())}")
    for an, am in action_model.items():
        print(f"  {an}: params={am['params']}, "
              f"pre={len(am['preconditions'])}, add={len(am['add_effects'])}, del={len(am['del_effects'])}")

    process_domain(args.base, args.output, action_model, title,
                   single_trajectory=args.single_trajectory)


if __name__ == "__main__":
    main()
